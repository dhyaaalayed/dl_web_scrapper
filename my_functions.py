import os
from pathlib import Path

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles import Protection


def get_template_columns(template_path, number_of_columns):
    def _get_hauschr_index(new_columns):
        for idx, column in enumerate(new_columns):
            if pd.isnull(column):
                return idx


    df = pd.read_excel(template_path)

    # Cut until columns row:
    df = df[5:]
    # Cut null columns:
    df = df.iloc[:, :number_of_columns]


    columns = list(df.iloc[0])
    print("template columns: ", columns)
    # set hauschar column
    hauschar_index = _get_hauschr_index(columns)
    columns[hauschar_index] = "Hauschar"

    return columns

def write_bvh_dfs_to_excel(path, bvh_city_name, df, bvh_installed_addresses_length):
    book = load_workbook(path)  # assuming that the new template is there
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    print("Log: Writing on Excel")

    sheet = book["HA_Auswertung"]
    sheet["A1"] = "All Montage: " + bvh_city_name
    sheet["AF4"] = bvh_installed_addresses_length

    fill = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')
    sheet["AF4"].fill = fill

    alignment = Alignment(horizontal='center', vertical='center')
    sheet["AF4"].alignment = alignment

    sheet.row_dimensions[7].protection = Protection(locked=True)


    df.to_excel(writer, index=False, startrow=7, startcol=0, sheet_name='HA_Auswertung', header=False)
    writer.save()


def parse_master_excel(excel_path, number_of_columns):
    """
        Takes path to master excel and the number_of_columns in the file!
            (because pandas is not able to discover the number of columns!)
        Returns a df of it
    """
    def get_hauschr_index(new_columns):
        for idx, column in enumerate(new_columns):
            print("idx: ", idx, " col: ", column)
            if pd.isnull(column):
                print("inside if: ", idx, " null col: ", column)
                return idx

    df = pd.read_excel(open(excel_path, 'rb'), sheet_name="HA_Auswertung")
    # then start with the 5th row, where the header of data is located
    df = df[5:]
    # set columns names
    new_columns = list(df.iloc[0])
    # Exception case (hauschar column): set a column name manually, because it's nan

    hauschar_index = get_hauschr_index(new_columns)
    new_columns[hauschar_index] = "Hauschar"
    df.columns = new_columns

    df = df.dropna(subset=['PLZ'])  # drop rows where all values are nan

    # Drop nan columns
    df = df.iloc[:, :number_of_columns]

    # Drop header row
    df = df[1:]
    df["Hauschar"] = df["Hauschar"].fillna("")  # fill na for hauschr for comparision
    df["Vorderhaus/Hinterhaus"] = df["Vorderhaus/Hinterhaus"].fillna("")
    return df

def log(message):
    print("Log: " + message)



def create_unique_id_for_master_df(df):
    """
        Takes a master df and create a unique_id column by concatenating the following columns:
        It has already been tested on Dresden rows and no more than row have the same id :)
    """
    df["unique_id"] = df["NVT"].apply(str) + "_" + df["PLZ"].apply(int).apply(str) + "_" + df["Ort"].apply(str) + "_" + df["Straße"].apply(str) + "_" + df["Hausnr."].apply(str) + "_" + df["Hauschar"].apply(str) + "_" + df["Vorderhaus/Hinterhaus"].apply(str)
    unique_column_list = list(df["unique_id"])
    print("unique_column_list: ", unique_column_list)
    log("Verifying that the generated unique_id column is unique!")
    # assert(len(unique_column_list) == len(set(unique_column_list)))
    return df

def get_unique_id_of_not_null_column(df, column_name):
    """
        Takes df
        Returns list of addresses id, where Kommentar Telekom is not null
        example:
            column_name = '          Kommentar Telekom'
    """
    series = df[df[column_name].notnull()].unique_id
    return list(series)

def get_old_column_data_for_master_list(old_df, new_df, column_name):
    old_column_not_null_ids = get_unique_id_of_not_null_column(old_df, column_name)
    log("Number of not null cells of column {} is {}".format(column_name, len(old_column_not_null_ids)))
    for id in old_column_not_null_ids:
        # assert(len(new_df[new_df.unique_id == id]) == 1)
        new_df.loc[new_df.unique_id == id, column_name] = old_df.loc[old_df.unique_id == id, column_name].to_string(index=False) # in order not to include the index in the returned string!
        print("old_df.loc[old_df.unique_id == id, column_name]: ", old_df.loc[old_df.unique_id == id, column_name])
        print("new_df.loc[new_df.unique_id == id, column_name]: ", new_df.loc[new_df.unique_id == id, column_name])
    return new_df




# telekom_kommentar_column_name = '          Kommentar Telekom'
# current_df = parse_master_excel("/Users/dlprojectsit/Desktop/Masterliste_Dresden.xlsx", 41)
# current_df = create_unique_id_for_master_df(current_df)
# #
# current_df.loc[current_df.unique_id == "42V1010_1157_Dresden_Weidentalstr._46_", telekom_kommentar_column_name].to_string()
# #
# current_df.loc[current_df.unique_id == "42V1010_1157_Dresden_Weidentalstr._46_", telekom_kommentar_column_name] = current_df.loc[current_df.unique_id == "42V1010_1157_Dresden_Weidentalstr._45_", telekom_kommentar_column_name].to_string(index=False)


# current_df.loc[current_df.unique_id == "42V1010_1157_Dresden_Weidentalstr._46_", telekom_kommentar_column_name]


